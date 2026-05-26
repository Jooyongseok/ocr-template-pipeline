"""OCR Template Pipeline - Hugging Face Spaces Gradio App.

GPU 가속 한글 손글씨 OCR. 템플릿 선택 → PDF 업로드 → 자동 OCR → 결과 검수 → 엑셀 다운로드.
"""
import os
import json
import tempfile
import shutil
from pathlib import Path
from dataclasses import dataclass, field

import gradio as gr
import numpy as np
import torch
from PIL import Image
from transformers import VisionEncoderDecoderModel, TrOCRProcessor

# ── OCR Engine ──

@dataclass
class OCRResult:
    text: str
    confidence: float
    candidates: list = field(default_factory=list)
    error: str = None


class TrOCREngine:
    def __init__(self, model_name="ddobokki/ko-trocr", device="cuda"):
        self.model_name = model_name
        self.device = device
        self.model = None
        self.processor = None

    def load(self):
        if self.model is not None:
            return
        print(f"Loading model: {self.model_name}")
        self.processor = TrOCRProcessor.from_pretrained(self.model_name)
        self.model = VisionEncoderDecoderModel.from_pretrained(self.model_name)
        try:
            self.model.to(self.device)
        except Exception:
            self.device = "cpu"
            self.model.to("cpu")
        self.model.eval()
        print(f"Model loaded on {self.device}")

    @torch.no_grad()
    def predict(self, images, max_new_tokens=64, num_beams=4):
        self.load()
        pixel_values = self.processor(images=images, return_tensors="pt").pixel_values.to(self.device)
        outputs = self.model.generate(
            pixel_values,
            max_new_tokens=max_new_tokens,
            num_beams=num_beams,
            return_dict_in_generate=True,
            output_scores=True,
        )
        texts = self.processor.batch_decode(outputs.sequences, skip_special_tokens=True)
        confs = []
        if hasattr(outputs, "sequences_scores") and outputs.sequences_scores is not None:
            for s in outputs.sequences_scores:
                confs.append(round(min(1.0, max(0.0, float(torch.exp(s)))), 4))
        else:
            confs = [0.5] * len(texts)
        return [OCRResult(text=t.strip(), confidence=c, candidates=[{"text": t.strip(), "confidence": c}]) for t, c in zip(texts, confs)]


# ── Global state ──
engine = TrOCREngine(
    model_name=os.environ.get("MODEL_NAME", "ddobokki/ko-trocr"),
    device="cuda" if torch.cuda.is_available() else "cpu",
)

TEMPLATES_DIR = Path(__file__).parent / "templates"
TEMPLATES_DIR.mkdir(exist_ok=True)

# Bundle agriculture template
AGRICULTURE_TEMPLATE = {
    "template_id": "agriculture_subsidy",
    "document_name": "기본직접지불금 지급대상자 등록신청서",
    "fields": [
        {"key": "applicant_name", "label": "등록신청인 성명", "field_type": "korean_name", "required": True, "bbox_norm": [0.25, 0.15, 0.15, 0.03], "group": "applicant"},
        {"key": "applicant_rrn", "label": "등록신청인 주민등록번호", "field_type": "resident_number", "required": True, "bbox_norm": [0.42, 0.15, 0.22, 0.03], "group": "applicant"},
        {"key": "applicant_account", "label": "계좌번호 은행명", "field_type": "account", "required": False, "bbox_norm": [0.25, 0.185, 0.35, 0.03], "group": "applicant"},
        {"key": "applicant_address", "label": "등록신청인 주소", "field_type": "address", "required": True, "bbox_norm": [0.25, 0.22, 0.55, 0.03], "group": "applicant"},
        {"key": "applicant_phone", "label": "전화번호", "field_type": "phone", "required": False, "bbox_norm": [0.65, 0.185, 0.2, 0.03], "group": "applicant"},
        {"key": "manager_name", "label": "경영주 성명", "field_type": "korean_name", "required": False, "bbox_norm": [0.25, 0.28, 0.15, 0.03], "group": "manager"},
        {"key": "manager_farmer_no", "label": "경영주 농업인번호", "field_type": "number_text", "required": False, "bbox_norm": [0.42, 0.28, 0.18, 0.03], "group": "manager"},
        {"key": "manager_address", "label": "경영주 주소", "field_type": "address", "required": False, "bbox_norm": [0.25, 0.315, 0.55, 0.03], "group": "manager"},
        {"key": "manager_phone", "label": "경영주 전화번호", "field_type": "phone", "required": False, "bbox_norm": [0.5, 0.35, 0.2, 0.03], "group": "manager"},
    ],
}

# Save default template
default_tpl_path = TEMPLATES_DIR / "agriculture_subsidy.json"
if not default_tpl_path.exists():
    default_tpl_path.write_text(json.dumps(AGRICULTURE_TEMPLATE, ensure_ascii=False, indent=2), encoding="utf-8")


def list_templates():
    tpls = []
    for f in sorted(TEMPLATES_DIR.glob("*.json")):
        data = json.loads(f.read_text("utf-8"))
        name = data.get("document_name", f.stem)
        count = len(data.get("fields", []))
        tpls.append(f"{name} ({count}fields) [{f.stem}]")
    return tpls


def load_template(name_str):
    stem = name_str.split("[")[-1].rstrip("]") if "[" in name_str else name_str
    path = TEMPLATES_DIR / f"{stem}.json"
    if path.exists():
        return json.loads(path.read_text("utf-8"))
    return None


# ── PDF -> Crops ──

def pdf_to_crops(pdf_path, template):
    import fitz
    doc = fitz.open(pdf_path)
    page = doc[0]
    pix = page.get_pixmap(dpi=150)
    page_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()

    crops = []
    for f in template.get("fields", []):
        bbox = f["bbox_norm"]
        x = int(bbox[0] * pix.width)
        y = int(bbox[1] * pix.height)
        w = int(bbox[2] * pix.width)
        h = int(bbox[3] * pix.height)
        x, y = max(0, x), max(0, y)
        w = min(w, pix.width - x)
        h = min(h, pix.height - y)
        if w < 2 or h < 2:
            crops.append({"field": f, "image": None})
            continue
        crop = page_img.crop((x, y, x + w, y + h))
        crops.append({"field": f, "image": crop})

    return page_img, crops


# ── Core OCR function ──

def run_ocr(pdf_file, template_name, progress=gr.Progress()):
    if pdf_file is None:
        return None, "PDF 파일을 업로드해주세요.", None, None

    template = load_template(template_name)
    if template is None:
        return None, "템플릿을 선택해주세요.", None, None

    progress(0.1, desc="PDF 분석 중...")
    page_img, crops = pdf_to_crops(pdf_file, template)

    # Filter valid crops
    valid = [(c["field"], c["image"]) for c in crops if c["image"] is not None]
    if not valid:
        return page_img, "추출 가능한 필드가 없습니다.", None, None

    progress(0.3, desc=f"OCR 추론 중 ({len(valid)}개 필드)...")
    images = [c[1] for c in valid]

    # Batch OCR
    batch_size = 16
    all_results = []
    for i in range(0, len(images), batch_size):
        batch = images[i:i + batch_size]
        results = engine.predict(batch)
        all_results.extend(results)
        progress(0.3 + 0.5 * (i + len(batch)) / len(images), desc=f"OCR 추론 중... ({i + len(batch)}/{len(images)})")

    progress(0.85, desc="결과 정리 중...")

    # Build results table
    rows = []
    results_data = []
    for (fld, _), result in zip(valid, all_results):
        conf_pct = f"{result.confidence * 100:.1f}%"
        status = "OK" if result.confidence >= 0.8 else ("검수필요" if result.confidence >= 0.5 else "오류")
        rows.append([
            fld.get("label", fld["key"]),
            fld.get("field_type", "text"),
            result.text,
            conf_pct,
            status,
            fld.get("group", ""),
        ])
        results_data.append({
            "field_key": fld["key"],
            "label": fld.get("label", fld["key"]),
            "field_type": fld.get("field_type", "text"),
            "value": result.text,
            "confidence": result.confidence,
            "status": status,
            "group": fld.get("group", ""),
            "required": fld.get("required", False),
        })

    # Save results JSON for download
    results_json = json.dumps(results_data, ensure_ascii=False, indent=2)
    json_path = tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w", encoding="utf-8")
    json_path.write(results_json)
    json_path.close()

    # Create Excel
    excel_path = create_excel(results_data, template)

    progress(1.0, desc="완료!")

    # Summary
    total = len(rows)
    ok_count = sum(1 for r in rows if r[4] == "OK")
    review_count = sum(1 for r in rows if r[4] == "검수필요")
    error_count = sum(1 for r in rows if r[4] == "오류")
    summary = f"총 {total}개 필드 | OK: {ok_count} | 검수필요: {review_count} | 오류: {error_count}"

    return page_img, summary, rows, excel_path


def create_excel(results_data, template):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    headers = ["필드명", "필드타입", "OCR결과", "신뢰도", "상태", "그룹", "필수여부"]
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    ok_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    review_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    error_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")

    for i, r in enumerate(results_data, 2):
        ws.cell(row=i, column=1, value=r["label"])
        ws.cell(row=i, column=2, value=r["field_type"])
        ws.cell(row=i, column=3, value=r["value"])
        ws.cell(row=i, column=4, value=f"{r['confidence']*100:.1f}%")
        status_cell = ws.cell(row=i, column=5, value=r["status"])
        ws.cell(row=i, column=6, value=r["group"])
        ws.cell(row=i, column=7, value="Y" if r["required"] else "")

        if r["status"] == "OK":
            status_cell.fill = ok_fill
        elif r["status"] == "검수필요":
            status_cell.fill = review_fill
        else:
            status_cell.fill = error_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path.close()
    wb.save(path.name)
    return path.name


# ── Template upload ──

def upload_template(file):
    if file is None:
        return "파일을 업로드해주세요.", gr.update()
    try:
        data = json.loads(Path(file).read_text("utf-8"))
        if "fields" not in data:
            return "유효하지 않은 템플릿 (fields 배열 필요)", gr.update()
        name = data.get("template_id", Path(file).stem)
        dest = TEMPLATES_DIR / f"{name}.json"
        shutil.copy(file, dest)
        return f"템플릿 '{name}' 저장 완료 ({len(data['fields'])}개 필드)", gr.update(choices=list_templates())
    except Exception as e:
        return f"오류: {e}", gr.update()


# ── Batch OCR ──

def run_batch_ocr(files, template_name, progress=gr.Progress()):
    if not files:
        return "PDF 파일을 업로드해주세요.", None, None

    template = load_template(template_name)
    if template is None:
        return "템플릿을 선택해주세요.", None, None

    all_rows = []
    all_results = []

    for idx, pdf_file in enumerate(files):
        progress((idx) / len(files), desc=f"문서 {idx+1}/{len(files)} 처리 중...")
        try:
            _, crops = pdf_to_crops(pdf_file, template)
            valid = [(c["field"], c["image"]) for c in crops if c["image"] is not None]
            if not valid:
                continue

            images = [c[1] for c in valid]
            results = engine.predict(images)

            doc_name = Path(pdf_file).stem
            for (fld, _), result in zip(valid, results):
                status = "OK" if result.confidence >= 0.8 else ("검수필요" if result.confidence >= 0.5 else "오류")
                all_rows.append([
                    doc_name,
                    fld.get("label", fld["key"]),
                    result.text,
                    f"{result.confidence*100:.1f}%",
                    status,
                ])
                all_results.append({
                    "document": doc_name,
                    "field_key": fld["key"],
                    "label": fld.get("label", fld["key"]),
                    "field_type": fld.get("field_type", "text"),
                    "value": result.text,
                    "confidence": result.confidence,
                    "status": status,
                    "group": fld.get("group", ""),
                    "required": fld.get("required", False),
                })
        except Exception as e:
            all_rows.append([Path(pdf_file).stem, "ERROR", str(e), "0%", "오류"])

    progress(1.0, desc="완료!")

    # Excel
    excel_path = create_batch_excel(all_results) if all_results else None

    total = len(all_rows)
    ok = sum(1 for r in all_rows if r[4] == "OK")
    summary = f"{len(files)}개 문서, {total}개 필드 | OK: {ok} | 문제: {total - ok}"

    return summary, all_rows, excel_path


def create_batch_excel(results_data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch OCR"

    headers = ["문서", "필드명", "필드타입", "OCR결과", "신뢰도", "상태", "그룹"]
    hfill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont

    for i, r in enumerate(results_data, 2):
        ws.cell(row=i, column=1, value=r.get("document", ""))
        ws.cell(row=i, column=2, value=r["label"])
        ws.cell(row=i, column=3, value=r["field_type"])
        ws.cell(row=i, column=4, value=r["value"])
        ws.cell(row=i, column=5, value=f"{r['confidence']*100:.1f}%")
        ws.cell(row=i, column=6, value=r["status"])
        ws.cell(row=i, column=7, value=r.get("group", ""))

    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

    path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path.close()
    wb.save(path.name)
    return path.name


# ── Gradio UI ──

THEME = gr.themes.Base(
    primary_hue=gr.themes.colors.indigo,
    neutral_hue=gr.themes.colors.zinc,
    font=gr.themes.GoogleFont("Noto Sans KR"),
).set(
    body_background_fill="#09090b",
    body_background_fill_dark="#09090b",
    block_background_fill="#18181b",
    block_background_fill_dark="#18181b",
    block_border_color="#3f3f46",
    block_border_color_dark="#3f3f46",
    input_background_fill="#27272a",
    input_background_fill_dark="#27272a",
)

with gr.Blocks(theme=THEME, title="OCR Pipeline", css="""
    .gradio-container { max-width: 1100px !important; }
    .status-ok { color: #22c55e; font-weight: 700; }
    .status-review { color: #f59e0b; font-weight: 700; }
    .status-error { color: #ef4444; font-weight: 700; }
""") as demo:

    gr.Markdown("""
    # OCR Template Pipeline v2
    **한글 손글씨 OCR 자동 추출 시스템** | 템플릿 선택 → PDF 업로드 → 자동 OCR → 엑셀 다운로드

    지원 필드: 이름, 주민번호, 전화번호, 주소, 날짜, 체크박스, 서명 등 12종
    """)

    with gr.Tabs():
        # ── Tab 1: Single OCR ──
        with gr.Tab("단건 OCR"):
            with gr.Row():
                with gr.Column(scale=1):
                    tpl_select = gr.Dropdown(
                        label="템플릿 선택",
                        choices=list_templates(),
                        value=list_templates()[0] if list_templates() else None,
                    )
                    pdf_input = gr.File(label="PDF 업로드", file_types=[".pdf"], type="filepath")
                    run_btn = gr.Button("OCR 실행", variant="primary", size="lg")

                with gr.Column(scale=2):
                    page_preview = gr.Image(label="문서 미리보기", height=400)
                    summary_text = gr.Textbox(label="결과 요약", interactive=False)

            results_table = gr.Dataframe(
                headers=["필드명", "필드타입", "OCR결과", "신뢰도", "상태", "그룹"],
                label="OCR 결과",
                interactive=True,
                wrap=True,
            )
            excel_download = gr.File(label="엑셀 다운로드")

            run_btn.click(
                fn=run_ocr,
                inputs=[pdf_input, tpl_select],
                outputs=[page_preview, summary_text, results_table, excel_download],
            )

        # ── Tab 2: Batch OCR ──
        with gr.Tab("배치 OCR (여러 장)"):
            with gr.Row():
                tpl_select_b = gr.Dropdown(
                    label="템플릿 선택",
                    choices=list_templates(),
                    value=list_templates()[0] if list_templates() else None,
                )
                batch_input = gr.File(label="PDF 여러 장 업로드", file_types=[".pdf"], file_count="multiple", type="filepath")
                batch_btn = gr.Button("일괄 OCR 실행", variant="primary", size="lg")

            batch_summary = gr.Textbox(label="배치 결과 요약", interactive=False)
            batch_table = gr.Dataframe(
                headers=["문서", "필드명", "OCR결과", "신뢰도", "상태"],
                label="배치 OCR 결과",
                interactive=True,
                wrap=True,
            )
            batch_excel = gr.File(label="배치 엑셀 다운로드")

            batch_btn.click(
                fn=run_batch_ocr,
                inputs=[batch_input, tpl_select_b],
                outputs=[batch_summary, batch_table, batch_excel],
            )

        # ── Tab 3: Template Management ──
        with gr.Tab("템플릿 관리"):
            gr.Markdown("""
            ### 템플릿 업로드
            JSON 형식의 템플릿 파일을 업로드하세요. 템플릿에는 `fields` 배열이 포함되어야 합니다.

            각 필드는 `key`, `label`, `field_type`, `bbox_norm` (정규화 좌표 [x, y, w, h])을 포함합니다.
            """)
            tpl_upload = gr.File(label="템플릿 JSON 업로드", file_types=[".json"], type="filepath")
            tpl_upload_btn = gr.Button("템플릿 저장")
            tpl_upload_result = gr.Textbox(label="결과", interactive=False)

            tpl_upload_btn.click(
                fn=upload_template,
                inputs=[tpl_upload],
                outputs=[tpl_upload_result, tpl_select],
            )

            gr.Markdown("### 현재 저장된 템플릿")
            tpl_list_text = gr.Textbox(
                value="\n".join(list_templates()) or "저장된 템플릿이 없습니다",
                label="템플릿 목록",
                interactive=False,
                lines=5,
            )

        # ── Tab 4: Info ──
        with gr.Tab("시스템 정보"):
            device_info = "GPU (CUDA)" if torch.cuda.is_available() else "CPU"
            if torch.cuda.is_available():
                device_info += f" - {torch.cuda.get_device_name(0)}"

            gr.Markdown(f"""
            ### 시스템 정보

            | 항목 | 값 |
            |------|------|
            | **OCR 모델** | {engine.model_name} |
            | **디바이스** | {device_info} |
            | **PyTorch** | {torch.__version__} |
            | **지원 필드** | 12종 (korean_name, resident_number, phone, address, date, ...) |

            ### CER (Character Error Rate)이란?

            OCR이 인식한 텍스트와 정답을 문자 단위로 비교하여, 잘못된 문자의 비율을 나타내는 지표입니다.

            **CER = (대체 + 누락 + 삽입) / 전체 문자 수**

            - Fine-tuned TrOCR-small: **CER 2.648%** (100글자 중 2.6글자만 오류)
            - ddobokki/ko-trocr: 범용 한국어 TrOCR (213M params)

            ### 향후 계획

            - 추출된 데이터를 기반으로 **통계 분석 + 시각화 대시보드** 생성
            - VLM (Qwen-VL, Varco) 모델 통합
            - Active Learning 자동 재학습
            """)

if __name__ == "__main__":
    demo.launch(server_name="0.0.0.0", server_port=7860)
