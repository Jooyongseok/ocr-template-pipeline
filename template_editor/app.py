"""OCR 템플릿 편집기 - 범용 문서 지원 Flask 서버.

사용자가 직접 PDF를 업로드하고, 바운딩 박스를 그려 필드를 정의한 후,
재사용 가능한 템플릿으로 저장/로드할 수 있다.
"""
import os
import io
import json
import base64
from pathlib import Path

import fitz  # PyMuPDF
from flask import Flask, render_template, request, jsonify

from field_presets import (
    FIELD_TYPES, get_all_presets, get_groups_and_colors,
    save_user_preset, delete_user_preset, load_user_presets,
    SAMPLE_PRESETS,
)

app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = BASE_DIR / "template"
TEMPLATE_DIR.mkdir(exist_ok=True)

_page_cache = {}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f:
        return jsonify(error="파일이 없습니다"), 400

    page_num = int(request.form.get("page", 1)) - 1
    data = f.read()
    filename = f.filename or "unknown"

    if filename.lower().endswith(".pdf"):
        doc = fitz.open(stream=data, filetype="pdf")
        if page_num >= len(doc):
            return jsonify(error=f"페이지 {page_num+1}이 없습니다 (총 {len(doc)}페이지)"), 400
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        img_bytes = pix.tobytes("png")
        total_pages = len(doc)
        doc.close()
    else:
        img_bytes = data
        total_pages = 1

    b64 = base64.b64encode(img_bytes).decode()
    _page_cache["current"] = img_bytes

    return jsonify(
        image=f"data:image/png;base64,{b64}",
        total_pages=total_pages,
        filename=filename,
    )


@app.route("/presets")
def presets():
    all_presets = get_all_presets()
    groups, colors = get_groups_and_colors()
    return jsonify(
        presets=all_presets,
        field_types=FIELD_TYPES,
        groups=groups,
        group_colors=colors,
    )


@app.route("/save-template", methods=["POST"])
def save_template():
    data = request.get_json()
    if not data:
        return jsonify(error="데이터가 없습니다"), 400

    template_id = data.get("template_id", "template_v1")
    filename = f"{template_id}.json"
    out_path = TEMPLATE_DIR / filename

    with open(out_path, "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)

    return jsonify(saved=str(out_path))


@app.route("/templates")
def list_templates():
    files = sorted(TEMPLATE_DIR.glob("*.json"))
    result = []
    for f in files:
        with open(f, encoding="utf-8") as fp:
            meta = json.load(fp)
        result.append({
            "filename": f.name,
            "template_id": meta.get("template_id", f.stem),
            "document_name": meta.get("document_name", ""),
            "field_count": len(meta.get("fields", [])),
        })
    return jsonify(templates=result)


@app.route("/load-template", methods=["POST"])
def load_template():
    data = request.get_json()
    filename = data.get("filename")
    if not filename:
        return jsonify(error="파일명이 없습니다"), 400

    path = TEMPLATE_DIR / filename
    if not path.exists():
        return jsonify(error="파일을 찾을 수 없습니다"), 404

    with open(path, encoding="utf-8") as fp:
        template = json.load(fp)
    return jsonify(template=template)


@app.route("/delete-template", methods=["POST"])
def delete_template():
    data = request.get_json()
    filename = data.get("filename")
    if not filename:
        return jsonify(error="파일명이 없습니다"), 400

    path = TEMPLATE_DIR / filename
    if path.exists():
        path.unlink()
        return jsonify(ok=True)
    return jsonify(error="파일을 찾을 수 없습니다"), 404


# ── 사용자 정의 프리셋 관리 ──

@app.route("/save-preset", methods=["POST"])
def save_preset():
    data = request.get_json()
    name = data.get("name", "")
    if not name:
        return jsonify(error="프리셋 이름이 필요합니다"), 400
    path = save_user_preset(name, data)
    return jsonify(ok=True, path=path)


@app.route("/delete-preset", methods=["POST"])
def delete_preset():
    data = request.get_json()
    name = data.get("name", "")
    if delete_user_preset(name):
        return jsonify(ok=True)
    return jsonify(error="프리셋을 찾을 수 없습니다"), 404


@app.route("/user-presets")
def user_presets():
    presets = load_user_presets()
    return jsonify(presets=presets)


# ── 엑셀 틀 업로드: 컬럼명 → 필드 제안 ──

@app.route("/upload-excel-schema", methods=["POST"])
def upload_excel_schema():
    """채워야 할 엑셀 파일을 업로드하면 시트별 컬럼명을 추출하여 필드 제안 목록을 반환한다.

    사용자는 이 목록을 기반으로 바운딩 박스를 그리고 타입을 선택한다.
    """
    f = request.files.get("file")
    if not f:
        return jsonify(error="파일이 없습니다"), 400

    filename = f.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        return jsonify(error="엑셀 파일(.xlsx, .xls)만 지원합니다"), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify(error=f"엑셀 파일 읽기 실패: {e}"), 400

    sheets = {}
    suggested_fields = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() if c else "" for c in row]
            break

        headers = [h for h in headers if h]
        sheets[sheet_name] = headers

        for col_name in headers:
            # 컬럼명에서 필드 타입 자동 추론
            field_type = _guess_field_type(col_name)
            key = _make_field_key(col_name)
            suggested_fields.append({
                "key": key,
                "label": col_name,
                "field_type": field_type,
                "required": False,
                "excel_sheet": sheet_name,
                "excel_column": col_name,
                "group": sheet_name,
                "bbox_norm": None,  # 사용자가 바운딩 박스를 직접 그려야 함
            })

    wb.close()

    return jsonify({
        "filename": filename,
        "sheets": sheets,
        "suggested_fields": suggested_fields,
        "total_fields": len(suggested_fields),
    })


def _guess_field_type(col_name: str) -> str:
    """컬럼명에서 필드 타입을 추론한다."""
    name = col_name.lower().replace(" ", "").replace("_", "")
    if any(k in name for k in ["주민등록", "주민번호", "rrn"]):
        return "resident_number"
    if any(k in name for k in ["전화", "연락처", "핸드폰", "휴대폰", "phone", "tel"]):
        return "phone"
    if any(k in name for k in ["계좌", "account"]):
        return "account"
    if any(k in name for k in ["주소", "address"]):
        return "address"
    if any(k in name for k in ["성명", "이름", "name"]):
        return "korean_name"
    if any(k in name for k in ["생년월일", "birth"]):
        return "date_or_birth"
    if any(k in name for k in ["날짜", "일자", "date"]):
        return "date"
    if any(k in name for k in ["관계", "relation"]):
        return "relation"
    if any(k in name for k in ["체크", "check", "확인"]):
        return "checkbox"
    if any(k in name for k in ["서명", "sign", "signature"]):
        return "signature"
    if any(k in name for k in ["번호", "no", "number"]):
        return "number_text"
    return "text"


def _make_field_key(col_name: str) -> str:
    """컬럼명을 필드 키로 변환 (영문+숫자+언더스코어)."""
    import re
    import unicodedata
    # 한글은 그대로 유지하되 공백/특수문자를 언더스코어로
    key = re.sub(r'[^\w가-힣]', '_', col_name)
    key = re.sub(r'_+', '_', key).strip('_').lower()
    return key[:50] if key else "field"


if __name__ == "__main__":
    import sys
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5000
    print(f"OCR 템플릿 편집기: http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, debug=True)
