"""최종 엑셀 출력 (계약서 10절)"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from validator import mask_rrn, mask_account


# 스타일
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
REVIEW_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
OK_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 10.2 applicants 시트 컬럼
APPLICANT_COLUMNS = [
    ("document_id", "문서ID"),
    ("source_pdf", "원본파일"),
    ("applicant_name", "등록신청인_성명"),
    ("applicant_rrn", "등록신청인_주민등록번호"),
    ("applicant_account", "등록신청인_계좌번호_은행명"),
    ("applicant_address", "등록신청인_주소"),
    ("applicant_phone", "등록신청인_전화번호"),
    ("manager_name", "경영주_성명"),
    ("manager_farmer_no", "경영주_농업인번호"),
    ("manager_application_type", "경영주_신청유형"),
    ("manager_address", "경영주_주소"),
    ("manager_village", "경영주_마을명"),
    ("manager_phone", "경영주_전화번호"),
    ("livestock_farm_checked", "축산농가_체크"),
    ("facility_farm_checked", "시설농가_체크"),
    ("other_farmer_name", "경영주외_성명"),
    ("other_farmer_birth", "경영주외_생년월일"),
    ("other_farmer_no", "경영주외_농업인번호"),
    ("other_farmer_relation", "경영주와의_관계"),
    ("family_info_confirm_checked", "가족관계_확인_체크"),
    ("document_status", "문서상태"),
    ("review_count", "검수필요건수"),
]


def write_excel(documents: list[dict], output_path: str, mask_pii: bool = True):
    """
    문서별 결과를 엑셀로 저장한다.

    Args:
        documents: [{"document_id", "source_pdf", "fields": {key: {...}}, "document_status", "review_count"}]
        output_path: 출력 엑셀 경로
        mask_pii: 개인정보 마스킹 여부
    """
    wb = Workbook()

    # --- applicants 시트 ---
    ws_app = wb.active
    ws_app.title = "applicants"
    _write_header(ws_app, [col[1] for col in APPLICANT_COLUMNS])

    for row_idx, doc in enumerate(documents, start=2):
        fields = doc.get("fields", {})
        for col_idx, (key, _) in enumerate(APPLICANT_COLUMNS, start=1):
            if key == "document_id":
                val = doc.get("document_id", "")
            elif key == "source_pdf":
                val = doc.get("source_pdf", "")
            elif key == "document_status":
                val = doc.get("document_status", "")
            elif key == "review_count":
                val = doc.get("review_count", 0)
            elif key in fields:
                f = fields[key]
                val = f.get("value", "")
                # 마스킹
                if mask_pii and val:
                    if f.get("field_type") == "resident_number":
                        val = mask_rrn(str(val))
                    elif f.get("field_type") == "account":
                        val = mask_account(str(val))
                # 상태 기반 셀 색상
                status = f.get("status", "")
                cell = ws_app.cell(row=row_idx, column=col_idx, value=val)
                if status in ("needs_review", "multiple_candidates", "low_confidence"):
                    cell.fill = REVIEW_FILL
                elif status in ("ocr_failed", "missing", "invalid_format"):
                    cell.fill = ERROR_FILL
                cell.border = THIN_BORDER
                continue
            else:
                val = ""
            cell = ws_app.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    _auto_width(ws_app)

    # --- family_members 시트 ---
    ws_fam = wb.create_sheet("family_members")
    fam_headers = ["document_id", "source_pdf", "family_group", "row_index", "side", "관계", "성명", "주민등록번호", "confidence_min", "status"]
    _write_header(ws_fam, fam_headers)

    fam_row = 2
    for doc in documents:
        fields = doc.get("fields", {})
        for r in range(1, 5):
            for side, side_label in [("left", "좌"), ("right", "우")]:
                rel_key = f"family_{r}_{side}_relation"
                name_key = f"family_{r}_{side}_name"
                rrn_key = f"family_{r}_{side}_rrn"
                rel = fields.get(rel_key, {})
                name = fields.get(name_key, {})
                rrn = fields.get(rrn_key, {})
                # 하나라도 값이 있으면 행 생성
                if any(f.get("value") for f in [rel, name, rrn]):
                    confs = [f.get("confidence", 0) for f in [rel, name, rrn] if f]
                    min_conf = min(confs) if confs else 0
                    statuses = [f.get("status", "") for f in [rel, name, rrn] if f]
                    worst = "ok"
                    for s in ["ocr_failed", "missing", "low_confidence", "needs_review"]:
                        if s in statuses:
                            worst = s
                            break

                    rrn_val = rrn.get("value", "")
                    if mask_pii and rrn_val:
                        rrn_val = mask_rrn(str(rrn_val))

                    values = [
                        doc.get("document_id", ""),
                        doc.get("source_pdf", ""),
                        "④-1",
                        r,
                        side_label,
                        rel.get("value", ""),
                        name.get("value", ""),
                        rrn_val,
                        round(min_conf, 4),
                        worst,
                    ]
                    for ci, v in enumerate(values, start=1):
                        cell = ws_fam.cell(row=fam_row, column=ci, value=v)
                        cell.border = THIN_BORDER
                    fam_row += 1
    _auto_width(ws_fam)

    # --- review_items 시트 ---
    ws_rev = wb.create_sheet("review_items")
    rev_headers = ["document_id", "source_pdf", "page", "field_key", "field_label", "field_type",
                   "raw_text", "value", "confidence", "status", "warning", "crop_path"]
    _write_header(ws_rev, rev_headers)

    rev_row = 2
    for doc in documents:
        fields = doc.get("fields", {})
        for key, f in fields.items():
            if f.get("status") not in ("ok", "unchecked"):
                values = [
                    doc.get("document_id", ""), doc.get("source_pdf", ""), f.get("page", 1),
                    key, f.get("label", ""), f.get("field_type", ""),
                    f.get("raw_text", ""), f.get("value", ""),
                    f.get("confidence", 0), f.get("status", ""), f.get("warning", ""),
                    f.get("crop_path", ""),
                ]
                for ci, v in enumerate(values, start=1):
                    cell = ws_rev.cell(row=rev_row, column=ci, value=v if v is not None else "")
                    cell.border = THIN_BORDER
                    if f.get("status") in ("ocr_failed", "missing"):
                        cell.fill = ERROR_FILL
                    else:
                        cell.fill = REVIEW_FILL
                rev_row += 1
    _auto_width(ws_rev)

    # --- raw_ocr_results 시트 ---
    ws_raw = wb.create_sheet("raw_ocr_results")
    raw_headers = ["request_id", "document_id", "page", "field_key", "field_type",
                   "text", "value", "confidence", "status", "error"]
    _write_header(ws_raw, raw_headers)

    raw_row = 2
    for doc in documents:
        for key, f in doc.get("fields", {}).items():
            values = [
                f.get("request_id", ""), doc.get("document_id", ""), f.get("page", 1),
                key, f.get("field_type", ""), f.get("raw_text", ""),
                str(f.get("value", "")), f.get("confidence", 0),
                f.get("status", ""), f.get("error", ""),
            ]
            for ci, v in enumerate(values, start=1):
                ws_raw.cell(row=raw_row, column=ci, value=v if v is not None else "").border = THIN_BORDER
            raw_row += 1
    _auto_width(ws_raw)

    wb.save(output_path)
    print(f"  엑셀 저장: {output_path}")


def _write_header(ws, headers):
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(40, max(8, max_len + 2))
