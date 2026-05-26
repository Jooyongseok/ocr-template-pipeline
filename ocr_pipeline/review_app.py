"""통합 검수 웹 앱 -- 인라인 수정, 연관 필드 그룹, Active Learning 수집.

기존 review_server.py를 대체한다.
"""
import base64
import os
import re

from flask import Flask, jsonify, render_template, request, send_file

from .data_store import DataStore
from .active_learning import ActiveLearningCollector
from .field_dependency import get_group_fields, get_field_group
from .excel_writer import write_excel
from .model_registry import load_config as load_model_config
import io

SAFE_DOC_ID = re.compile(r"^[a-zA-Z0-9_\-]+$")


def create_app(work_dir: str = "work", output_dir: str = "output") -> Flask:
    app = Flask(
        __name__,
        template_folder=os.path.join(os.path.dirname(__file__), "templates"),
        static_folder=os.path.join(os.path.dirname(__file__), "static"),
    )

    store = DataStore(work_dir)
    collector = ActiveLearningCollector(work_dir)

    # ── 페이지 ──

    @app.route("/")
    def index():
        return render_template("dashboard.html")

    @app.route("/review")
    def review_page():
        return render_template("review.html")

    # ── API: PDF 업로드 (단건/배치) ──

    @app.route("/api/upload-pdf", methods=["POST"])
    def api_upload_pdf():
        """PDF 파일을 업로드하여 input/ 디렉토리에 저장한다."""
        files = request.files.getlist("files")
        if not files:
            return jsonify({"error": "파일이 없습니다"}), 400

        input_dir = os.path.join(os.path.dirname(work_dir), "input")
        os.makedirs(input_dir, exist_ok=True)

        saved = []
        for f in files:
            if f.filename and f.filename.lower().endswith(".pdf"):
                safe_name = re.sub(r'[^\w\-.]', '_', f.filename)
                path = os.path.join(input_dir, safe_name)
                f.save(path)
                saved.append(safe_name)

        return jsonify({"ok": True, "saved": saved, "count": len(saved)})

    # ── API: OCR 파이프라인 실행 ──

    @app.route("/api/run-ocr", methods=["POST"])
    def api_run_ocr():
        """선택된 템플릿과 PDF로 OCR 파이프라인을 실행한다."""
        data = request.get_json()
        if not data:
            return jsonify({"error": "요청 데이터가 없습니다"}), 400
        template_file = data.get("template", "")
        pdf_files = data.get("pdf_files", [])  # 빈 배열이면 input/ 전체
        user_id = data.get("user_id", "default")

        if not template_file or not template_file.endswith(".json"):
            return jsonify({"error": "템플릿 파일을 선택하세요"}), 400

        template_dir = os.path.join(os.path.dirname(work_dir), "template")
        template_path = os.path.join(template_dir, template_file)
        if not os.path.isfile(template_path):
            return jsonify({"error": f"템플릿 '{template_file}'을 찾을 수 없습니다"}), 400

        input_dir = os.path.join(os.path.dirname(work_dir), "input")

        # 대상 PDF 결정
        if pdf_files:
            targets = [os.path.join(input_dir, f) for f in pdf_files if os.path.exists(os.path.join(input_dir, f))]
        else:
            import glob
            targets = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))

        if not targets:
            return jsonify({"error": "처리할 PDF 파일이 없습니다"}), 400

        # 사용자별 작업 디렉토리
        user_work = os.path.join(work_dir, user_id) if user_id != "default" else work_dir
        os.makedirs(user_work, exist_ok=True)

        return jsonify({
            "ok": True,
            "template": template_file,
            "pdf_count": len(targets),
            "pdfs": [os.path.basename(t) for t in targets],
            "user_id": user_id,
            "work_dir": user_work,
            "message": f"{len(targets)}개 PDF에 대해 OCR을 실행합니다. CLI에서 실행하세요:\n"
                       f"python -m ocr_pipeline.run_pipeline --template {template_path} --input {input_dir} --work-dir {user_work}",
        })

    # ── API: 입력 PDF 목록 ──

    @app.route("/api/input-pdfs")
    def api_input_pdfs():
        input_dir = os.path.join(os.path.dirname(work_dir), "input")
        if not os.path.exists(input_dir):
            return jsonify({"pdfs": []})

        import glob
        pdfs = []
        for f in sorted(glob.glob(os.path.join(input_dir, "*.pdf"))):
            pdfs.append({
                "filename": os.path.basename(f),
                "size_kb": round(os.path.getsize(f) / 1024, 1),
            })
        return jsonify({"pdfs": pdfs})

    # ── API: 엑셀 스키마 업로드 (필드 추출) ──

    @app.route("/api/upload-excel-schema", methods=["POST"])
    def api_upload_excel_schema():
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "파일이 없습니다"}), 400
        filename = f.filename or ""
        if not filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"error": "엑셀 파일(.xlsx)만 지원합니다"}), 400
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        except Exception as e:
            return jsonify({"error": f"엑셀 읽기 실패: {e}"}), 400

        sheets = {}
        suggested = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c).strip() for c in row if c]
                break
            sheets[sn] = headers
            for col in headers:
                suggested.append({"key": col, "label": col, "excel_sheet": sn, "excel_column": col})
        wb.close()
        return jsonify({"filename": filename, "sheets": sheets, "suggested_fields": suggested, "total_fields": len(suggested)})

    # ── API: 문서 목록 ──

    @app.route("/api/documents")
    def api_documents():
        doc_ids = store.list_documents()
        docs = []
        for did in doc_ids:
            doc = store.get_document(did)
            if doc:
                docs.append({
                    "document_id": did,
                    "source_pdf": doc.get("source_pdf", ""),
                    "status": doc.get("document_status", "unknown"),
                    "review_count": doc.get("review_count", 0),
                    "total_fields": len(doc.get("fields", {})),
                    "dependency_warnings": doc.get("dependency_warnings", []),
                })
        return jsonify(docs)

    # ── API: 검수 대상 필드 (페이지네이션) ──

    @app.route("/api/review-items")
    def api_review_items():
        doc_id = request.args.get("doc_id")
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 50))

        items = store.get_review_items(doc_id)
        total = len(items)
        start = (page - 1) * per_page
        end = start + per_page

        return jsonify({
            "items": items[start:end],
            "total": total,
            "page": page,
            "per_page": per_page,
            "has_next": end < total,
        })

    # ── API: 문서 전체 필드 (검수 화면용) ──

    @app.route("/api/document/<doc_id>/fields")
    def api_document_fields(doc_id):
        if not SAFE_DOC_ID.match(doc_id):
            return jsonify({"error": "잘못된 document_id"}), 400

        doc = store.get_document(doc_id)
        if doc is None:
            return jsonify({"error": "문서 없음"}), 404

        fields = []
        for fk, field in doc.get("fields", {}).items():
            group = get_field_group(fk)
            fields.append({
                "field_key": fk,
                "label": field.get("label", fk),
                "field_type": field.get("field_type", "text"),
                "value": field.get("value", ""),
                "raw_text": field.get("raw_text", ""),
                "confidence": field.get("confidence", 0.0),
                "status": field.get("status", "ok"),
                "warning": field.get("warning"),
                "candidates": field.get("candidates", []),
                "bbox_norm": field.get("bbox_norm"),
                "crop_path": field.get("crop_path", ""),
                "edited": field.get("edited", False),
                "group_name": group["group_name"] if group else None,
                "group_fields": get_group_fields(fk),
                "required": field.get("required", False),
            })

        # 신뢰도 오름차순 (worst-first), 문제 필드 먼저
        fields.sort(key=lambda x: (
            0 if x["status"] not in ("ok", "unchecked") else 1,
            x["confidence"],
        ))

        return jsonify({
            "document_id": doc_id,
            "source_pdf": doc.get("source_pdf", ""),
            "document_status": doc.get("document_status", "unknown"),
            "review_count": doc.get("review_count", 0),
            "dependency_warnings": doc.get("dependency_warnings", []),
            "fields": fields,
        })

    # ── API: crop 이미지 (lazy-loading) ──

    @app.route("/api/crop/<doc_id>/<field_key>")
    def api_crop(doc_id, field_key):
        if not SAFE_DOC_ID.match(doc_id):
            return jsonify({"error": "잘못된 document_id"}), 400

        doc = store.get_document(doc_id)
        if doc is None:
            return jsonify({"error": "문서 없음"}), 404

        field = doc.get("fields", {}).get(field_key)
        if field is None:
            return jsonify({"error": "필드 없음"}), 404

        crop_path = field.get("crop_path", "")
        if not crop_path or not os.path.exists(crop_path):
            return jsonify({"error": "crop 이미지 없음"}), 404

        # 보안: crop_path가 work_dir 내부인지 확인
        abs_crop = os.path.abspath(crop_path)
        abs_work = os.path.abspath(work_dir)
        if not abs_crop.startswith(abs_work):
            return jsonify({"error": "접근 거부"}), 403

        return send_file(abs_crop, mimetype="image/png")

    # ── API: 페이지 이미지 ──

    @app.route("/api/page-image/<doc_id>/<int:page_num>")
    def api_page_image(doc_id, page_num):
        if not SAFE_DOC_ID.match(doc_id):
            return jsonify({"error": "잘못된 document_id"}), 400

        page_path = os.path.join(work_dir, "page_images", doc_id, f"page_{page_num:03d}.png")
        if not os.path.exists(page_path):
            return jsonify({"error": "페이지 이미지 없음"}), 404

        abs_page = os.path.abspath(page_path)
        abs_work = os.path.abspath(work_dir)
        if not abs_page.startswith(abs_work):
            return jsonify({"error": "접근 거부"}), 403

        return send_file(abs_page, mimetype="image/png")

    # ── API: 필드 수정 ──

    @app.route("/api/update", methods=["POST"])
    def api_update():
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"error": "요청 데이터 없음"}), 400

        doc_id = data.get("document_id", "")
        field_key = data.get("field_key", "")
        new_value = data.get("value", "")

        if not SAFE_DOC_ID.match(doc_id):
            return jsonify({"error": "잘못된 document_id"}), 400

        try:
            # 수정 전 원본 정보 (Active Learning용)
            doc_before = store.get_document(doc_id)
            field_before = doc_before.get("fields", {}).get(field_key, {}) if doc_before else {}
            original_text = field_before.get("raw_text", field_before.get("value", ""))
            original_confidence = field_before.get("confidence", 0.0)
            crop_path = field_before.get("crop_path", "")
            field_type = field_before.get("field_type", "text")

            # 값 저장
            updated_doc = store.save_field(doc_id, field_key, new_value)

            # Active Learning: 수정 데이터 수집
            al_result = {}
            if original_text != new_value and crop_path:
                al_result = collector.collect_correction(
                    crop_path=crop_path,
                    corrected_text=new_value,
                    field_type=field_type,
                    original_text=original_text,
                    original_confidence=original_confidence,
                    document_id=doc_id,
                    field_key=field_key,
                )

            return jsonify({
                "ok": True,
                "document_status": updated_doc.get("document_status"),
                "review_count": updated_doc.get("review_count"),
                "dependency_warnings": updated_doc.get("dependency_warnings", []),
                "active_learning": al_result,
            })
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

    # ── API: 문서 건너뛰기 ──

    @app.route("/api/skip-document", methods=["POST"])
    def api_skip_document():
        data = request.get_json()
        doc_id = data.get("document_id", "")
        reason = data.get("reason", "skipped")

        if not SAFE_DOC_ID.match(doc_id):
            return jsonify({"error": "잘못된 document_id"}), 400

        doc = store.get_document(doc_id)
        if doc is None:
            return jsonify({"error": "문서 없음"}), 404

        doc["document_status"] = "skipped"
        doc["skip_reason"] = reason
        store._atomic_write(store._doc_path(doc_id), doc)

        return jsonify({"ok": True, "document_status": "skipped"})

    # ── API: 통계 ──

    @app.route("/api/stats")
    def api_stats():
        batch_stats = store.get_batch_stats()
        al_stats = collector.get_stats()
        return jsonify({
            **batch_stats,
            "active_learning": al_stats,
        })

    # ── API: 엑셀 내보내기 ──

    @app.route("/api/export-excel", methods=["POST"])
    def api_export_excel():
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "result_reviewed.xlsx")

        docs = []
        for did in store.list_documents():
            doc = store.get_document(did)
            if doc:
                docs.append(doc)

        if not docs:
            return jsonify({"error": "내보낼 문서 없음"}), 400

        write_excel(docs, output_path, mask_pii=True)
        return jsonify({"ok": True, "path": output_path})

    # ── API: Active Learning 통계 ──

    @app.route("/api/learning-stats")
    def api_learning_stats():
        stats = collector.get_stats()
        dist = collector.get_confidence_distribution()
        return jsonify({
            **stats,
            "confidence_distribution": dist,
            "trigger_count": collector.trigger_count,
            "ready_for_retrain": stats.get("total_corrections", 0) >= collector.trigger_count,
        })

    # ── API: Active Learning 데이터 내보내기 ──

    @app.route("/api/export-training-data", methods=["POST"])
    def api_export_training():
        path = collector.export_for_training()
        if not path:
            return jsonify({"error": "수집된 데이터 없음"}), 400
        return jsonify({"ok": True, "path": path, "stats": collector.get_stats()})

    # ── API: 모델 목록 ──

    @app.route("/api/models")
    def api_models():
        config = load_model_config()
        models_conf = config.get("models", {})
        default_model = config.get("default_model", "")
        result = []
        for mid, conf in models_conf.items():
            result.append({
                "id": mid,
                "description": conf.get("description", mid),
                "engine": conf.get("engine", "unknown"),
                "is_default": mid == default_model,
            })
        return jsonify({"models": result, "current": default_model})

    # ── API: 모델 전환 ──

    @app.route("/api/switch-model", methods=["POST"])
    def api_switch_model():
        data = request.get_json()
        model_id = data.get("model_id", "")
        config = load_model_config()
        models_conf = config.get("models", {})
        if model_id not in models_conf:
            return jsonify({"error": f"모델 '{model_id}'을 찾을 수 없습니다"}), 400
        # Update config file
        import yaml
        config["default_model"] = model_id
        config_path = os.path.join(os.path.dirname(__file__), "model_config.yaml")
        with open(config_path, "w", encoding="utf-8") as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False)
        return jsonify({"ok": True, "current": model_id})

    # ── API: 원격 수정 데이터 수신 (다른 컴퓨터에서 보내는 교정 데이터) ──

    @app.route("/api/sync-corrections", methods=["POST"])
    def api_sync_corrections():
        """다른 사용자/컴퓨터에서 전송된 교정 데이터를 수신하여 저장한다.

        요청 형식:
        {
            "corrections": [
                {
                    "document_id": "...",
                    "field_key": "...",
                    "original_text": "...",
                    "corrected_text": "...",
                    "original_confidence": 0.45,
                    "corrected_by": "user@remote",
                    "corrected_at": "2026-05-26T12:00:00"
                }, ...
            ]
        }
        """
        data = request.get_json(force=True, silent=True)
        if not data or "corrections" not in data:
            return jsonify({"error": "corrections 배열이 필요합니다"}), 400

        corrections = data["corrections"]
        sync_dir = os.path.join(work_dir, "sync_corrections")
        os.makedirs(sync_dir, exist_ok=True)

        import json as _json
        from datetime import datetime

        sync_file = os.path.join(sync_dir, "received_corrections.jsonl")
        saved = 0
        for corr in corrections:
            if not corr.get("document_id") or not corr.get("field_key"):
                continue
            corr["received_at"] = datetime.now().isoformat()
            with open(sync_file, "a", encoding="utf-8") as f:
                f.write(_json.dumps(corr, ensure_ascii=False) + "\n")

            # 해당 문서의 필드 값도 업데이트
            doc_id = corr["document_id"]
            if SAFE_DOC_ID.match(doc_id):
                try:
                    store.save_field(doc_id, corr["field_key"], corr["corrected_text"])
                    saved += 1
                except Exception:
                    pass

        return jsonify({"ok": True, "saved": saved, "total": len(corrections)})

    # ── API: 교정 데이터 내보내기 (git push용) ──

    @app.route("/api/export-corrections")
    def api_export_corrections():
        """로컬에서 수정된 교정 데이터를 JSON으로 내보낸다."""
        corrections = store.get_all_corrections()
        return jsonify({"corrections": corrections, "total": len(corrections)})

    # ── API: 템플릿 목록 ──

    @app.route("/api/templates")
    def api_templates():
        template_dir = os.path.join(os.path.dirname(work_dir), "template")
        if not os.path.exists(template_dir):
            return jsonify({"templates": []})

        templates = []
        import glob
        for f in sorted(glob.glob(os.path.join(template_dir, "*.json"))):
            import json as _json
            with open(f, encoding="utf-8") as fp:
                try:
                    meta = _json.load(fp)
                    templates.append({
                        "filename": os.path.basename(f),
                        "template_id": meta.get("template_id", os.path.splitext(os.path.basename(f))[0]),
                        "document_name": meta.get("document_name", ""),
                        "field_count": len(meta.get("fields", [])),
                    })
                except _json.JSONDecodeError:
                    pass
        return jsonify({"templates": templates})

    return app


def main():
    import argparse
    parser = argparse.ArgumentParser(description="OCR 검수 웹 앱")
    parser.add_argument("--work-dir", default="work")
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=5001)
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    app = create_app(args.work_dir, args.output_dir)
    print(f"\n  검수 앱 실행: http://{args.host}:{args.port}")
    print(f"  작업 디렉토리: {args.work_dir}")
    print(f"  출력 디렉토리: {args.output_dir}\n")
    app.run(host=args.host, port=args.port, debug=args.debug)


if __name__ == "__main__":
    main()
